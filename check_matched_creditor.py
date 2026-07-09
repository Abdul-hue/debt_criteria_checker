
from openpyxl import load_workbook

# Load the workbook
wb = load_workbook('reconciliation_output/matched_creditor.xlsx')
ws = wb.active

# Get headers
headers = [cell.value for cell in ws[1]]
print("Headers in file:")
for i, header in enumerate(headers):
    print(f"{i}: {header}")
