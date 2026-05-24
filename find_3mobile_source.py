import os
import django
import sys
from pathlib import Path
import openpyxl

# Setup Django
sys.path.append(str(Path(__file__).resolve().parent))
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'debt_project.settings')
django.setup()

from debt_app.models import CreditorCriteria

def find_exact_location():
    excel_path = r"C:\Users\Canton Computers\Desktop\TIP CRITERIA & VOTING HISTORY.xlsx"
    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    
    target = "3 Mobile"
    found = False
    
    print(f"Searching for exact string '{target}' in all sheets...")
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            for col_idx, value in enumerate(row, start=1):
                if value and str(value).strip() == target:
                    print(f"FOUND in sheet '{sheet_name}' at Row {row_idx}, Col {col_idx}")
                    found = True
    
    if not found:
        print(f"Exact string '{target}' NOT found in any sheet.")
        
    # Also check what's in the DB for that name
    db_entries = CreditorCriteria.objects.filter(creditor_name=target)
    if db_entries.exists():
        print(f"\nDB has {db_entries.count()} entry for '{target}':")
        for e in db_entries:
            print(f"  ID: {e.id}, Source: {e.source_sheet}, Rep: {e.representative}")
    else:
        print(f"\nDB has NO entry for '{target}'")

if __name__ == "__main__":
    find_exact_location()
