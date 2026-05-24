"""
Extract data from TIP CRITERIA & VOTING HISTORY.xlsx
Sheets: Dividends, Banking Groups, County Councils, Which Representative
Output: Separate .md files in the Excel Criteria directory
"""

import openpyxl
from pathlib import Path

EXCEL_FILE = r"C:\Users\Canton Computers\Desktop\TIP CRITERIA & VOTING HISTORY.xlsx"
OUTPUT_DIR = Path(r"C:\Users\Canton Computers\Desktop\Debt Criteria check\Excel Criteria")

# --------------------------------------------------------------------------- #
# Helpers
# --------------------------------------------------------------------------- #

def clean(val):
    """Return a clean string from a cell value."""
    if val is None:
        return ""
    s = str(val).replace("\xa0", "").replace("\n", " ").strip()
    return s


def md_row(cells):
    """Return a markdown table row string."""
    return "| " + " | ".join(cells) + " |"


def md_separator(n_cols):
    return "| " + " | ".join(["---"] * n_cols) + " |"


# --------------------------------------------------------------------------- #
# Sheet extractors
# --------------------------------------------------------------------------- #

def extract_dividends(wb):
    ws = wb["Dividends "]
    rows = list(ws.iter_rows(values_only=True))

    lines = ["# Dividends", ""]
    headers = [clean(h) or "—" for h in rows[0]]
    lines.append(md_row(headers))
    lines.append(md_separator(len(headers)))

    for row in rows[1:]:
        cells = [clean(v) for v in row]
        if any(cells):           # skip entirely blank rows
            lines.append(md_row(cells))

    lines.append("")
    return "\n".join(lines)


def extract_banking_groups(wb):
    ws = wb["Banking Groups"]
    rows = list(ws.iter_rows(values_only=True))

    # Discover column groups – the sheet has group headers in row 0
    # Layout: GroupName | blank | GroupName | blank | GroupName …
    lines = ["# Banking Groups", ""]

    # Collect group header names from first row (non-None, non-blank columns)
    group_headers = []
    col_indices = []
    for i, val in enumerate(rows[0]):
        c = clean(val)
        if c:
            group_headers.append(c)
            col_indices.append(i)

    # Build per-group lists (each group spans from col_indices[n] to col_indices[n+1]-1)
    groups = {h: [] for h in group_headers}
    for row in rows[2:]:          # row 1 is blank spacer
        for idx, header in zip(col_indices, group_headers):
            val = clean(row[idx]) if idx < len(row) else ""
            if val:
                groups[header].append(val)

    # Output each group as a section
    for header in group_headers:
        lines.append(f"## {header}")
        lines.append("")
        for member in groups[header]:
            lines.append(f"- {member}")
        lines.append("")

    return "\n".join(lines)


def extract_county_councils(wb):
    ws = wb["County Councils"]
    rows = list(ws.iter_rows(values_only=True))

    lines = ["# County Councils", ""]

    # Row 0 = headers; columns: Council | Accept/Reject | Notes | Districts | (merged)
    # We'll use a card-style format because Districts text is very long
    data_rows = rows[1:]
    for row in data_rows:
        council      = clean(row[0]) if len(row) > 0 else ""
        accept       = clean(row[1]) if len(row) > 1 else ""
        notes        = clean(row[2]) if len(row) > 2 else ""
        districts_raw = clean(row[3]) if len(row) > 3 else ""

        if not council:
            continue

        lines.append(f"## {council}")
        lines.append("")
        if accept:
            lines.append(f"**Accept / Reject:** {accept}  ")
        if notes:
            lines.append(f"**Notes:** {notes}  ")
        if districts_raw:
            lines.append("")
            lines.append("**Districts:**")
            # Split on multiple whitespace or dashes that act as separators
            import re
            district_entries = re.split(r'\s{3,}', districts_raw)
            for d in district_entries:
                d = d.strip(" -–")
                if d:
                    lines.append(f"- {d}")
        lines.append("")

    return "\n".join(lines)


def extract_which_representative(wb):
    ws = wb["Which Representative"]
    rows = list(ws.iter_rows(values_only=True))

    lines = ["# Which Representative", ""]

    # Row 0 = column group headers (e.g. TIX, (WATCH) WPM, EVOLVE, EVERYDAY LOANS)
    # Columns are in pairs: creditor name | blank | creditor name | blank …
    # Detect header columns (even-indexed non-blank values in row 0)
    header_row = rows[0]
    group_headers = []
    col_indices = []
    for i, val in enumerate(header_row):
        c = clean(val)
        if c:
            group_headers.append(c)
            col_indices.append(i)

    # Build per-group creditor lists
    groups = {h: [] for h in group_headers}
    for row in rows[1:]:
        for idx, header in zip(col_indices, group_headers):
            val = clean(row[idx]) if idx < len(row) else ""
            if val:
                groups[header].append(val)

    # Notes row (row 0, last group sometimes contains a note in the value cell)
    # Already captured via groups

    for header in group_headers:
        lines.append(f"## {header}")
        lines.append("")
        creditors = groups[header]
        if creditors:
            for c in creditors:
                lines.append(f"- {c}")
        lines.append("")

    return "\n".join(lines)


# --------------------------------------------------------------------------- #
# Main
# --------------------------------------------------------------------------- #

def main():
    print(f"Loading workbook: {EXCEL_FILE}")
    wb = openpyxl.load_workbook(EXCEL_FILE, read_only=True)

    tasks = [
        ("Dividends ",        "Dividends_Criteria.md",           extract_dividends),
        ("Banking Groups",    "Banking_Groups_Criteria.md",       extract_banking_groups),
        ("County Councils",   "County_Councils_Criteria.md",      extract_county_councils),
        ("Which Representative", "Which_Representative_Criteria.md", extract_which_representative),
    ]

    for sheet_name, filename, extractor in tasks:
        print(f"  Extracting sheet: '{sheet_name}' -> {filename}")
        content = extractor(wb)
        out_path = OUTPUT_DIR / filename
        out_path.write_text(content, encoding="utf-8")
        print(f"    - Saved {out_path}")

    wb.close()
    print("\nAll sheets extracted successfully.")


if __name__ == "__main__":
    main()
